import pandas as pd
import os
from pathlib import Path
import openpyxl
from pypdf import PdfReader

def load_text(folder_path: str):
    documents = []
    folder = Path(folder_path)
    for file in folder.glob("*.txt"):
        with open(file, "r", encoding="utf-8") as f:
            content = f.read()
            documents.append({
                "source": file.name,
                "content": content
            })
    print(f"✅ Loaded {len(documents)} text documents from {folder}")
    return documents

def load_xlsx(file_path):
    df = pd.read_excel(file_path)

    chunks = []
    for idx, row in df.iterrows():
        row_text = ". ".join(
            f"{col}: {row[col]}"
            for col in df.columns
            if pd.notna(row[col])
        )

        chunks.append({
            "source": str(file_path),
            "content": row_text
        })
    print(f"✅ Loaded {len(chunks)} rows from {file_path}")
    return chunks

def load_documents(data_dir: str):
    """
    Load documents from data directory.
    Supports: .txt, .xlsx, .pdf
    """
    documents = []
    data_path = Path(data_dir)

    if not data_path.exists():
        print(f"⚠️ Data directory '{data_dir}' not found. Creating...")
        data_path.mkdir(parents=True, exist_ok=True)
        return documents

    # ---- Load .txt files ----
    for txt_file in data_path.glob("*.txt"):
        with open(txt_file, "r", encoding="utf-8") as f:
            content = f.read()
            documents.append({
                "source": txt_file.name,
                "content": content
            })
        print(f"✅ Loaded: {txt_file.name}")

    # ---- Load .xlsx files ----
    for xlsx_file in data_path.glob("*.xlsx"):
        workbook = openpyxl.load_workbook(xlsx_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content = "\n".join(
                " | ".join(str(cell.value) for cell in row if cell.value)
                for row in sheet.iter_rows()
            )
            documents.append({
                "source": f"{xlsx_file.name}#{sheet_name}",
                "content": content
            })
        print(f"✅ Loaded: {xlsx_file.name}")

    # ---- Load .pdf files (NEW) ----
    for pdf_file in data_path.glob("*.pdf"):
        try:
            with open(pdf_file, "rb") as f:
                pdf_reader = PdfReader(f)
                
                # Extract text from all pages
                full_text = ""
                for page_num, page in enumerate(pdf_reader.pages):
                    text = page.extract_text()
                    full_text += f"\n[Page {page_num + 1}]\n{text}"
                
                documents.append({
                    "source": pdf_file.name,
                    "content": full_text
                })
            print(f"✅ Loaded: {pdf_file.name}")
        except Exception as e:
            print(f"❌ Error loading {pdf_file.name}: {e}")

    return documents

def load_single_file(file_path: str):
    file_path = Path(file_path)
    if not file_path.exists():
        print(f"⚠️ File '{file_path}' not found.")
        return None

    if file_path.suffix == ".txt":
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
            print(f"✅ Loaded: {file_path.name}")
            return [{
                "source": file_path.name,
                "content": content
            }]
    elif file_path.suffix == ".xlsx":
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        content = "\n".join(
            " | ".join(str(cell.value) for cell in row if cell.value)
            for row in sheet.iter_rows()
        )
        print(f"✅ Loaded: {file_path.name}")
        return [{
            "source": file_path.name,
            "content": content
        }]
    elif file_path.suffix == ".pdf":
        try:
            with open(file_path, "rb") as f:
                pdf_reader = PdfReader(f)
                full_text = ""
                for page_num, page in enumerate(pdf_reader.pages):
                    text = page.extract_text()
                    full_text += f"\n[Page {page_num + 1}]\n{text}"
                print(f"✅ Loaded: {file_path.name}")
                return [{
                    "source": file_path.name,
                    "content": full_text
                }]
        except Exception as e:
            print(f"❌ Error loading {file_path.name}: {e}")
            return None
    else:
        print(f"⚠️ Unsupported file type: {file_path.suffix}")
        return None
